import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, date, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import pandas as pd
from utils.analytics import LabAnalytics

class AnalyticsDashboard:
    def __init__(self, parent, database):
        self.parent = parent
        self.database = database
        self.analytics = LabAnalytics(database)
        
        # Create main frame
        self.main_frame = ttk.Frame(parent)
        self.main_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        # Configure grid weights
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(1, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)
        
        self.setup_dashboard()
    
    def setup_dashboard(self):
        """Setup the analytics dashboard"""
        # Title
        title_label = ttk.Label(self.main_frame, text="Analytics Dashboard", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Create notebook for different analytics sections
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 10))
        
        # Create tabs
        self.create_overview_tab()
        self.create_revenue_tab()
        self.create_tests_tab()
        self.create_patients_tab()
        self.create_outstanding_tab()
        
        # Control panel
        self.create_control_panel()
        
        # Load initial data
        self.refresh_data()
    
    def create_control_panel(self):
        """Create control panel with refresh and export options"""
        control_frame = ttk.LabelFrame(self.main_frame, text="Controls", padding="10")
        control_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        
        # Date range selection
        ttk.Label(control_frame, text="From:").grid(row=0, column=0, padx=(0, 5))
        self.start_date_var = tk.StringVar(value=(date.today().replace(day=1)).strftime('%Y-%m-%d'))
        self.start_date_entry = ttk.Entry(control_frame, textvariable=self.start_date_var, width=12)
        self.start_date_entry.grid(row=0, column=1, padx=(0, 10))
        
        ttk.Label(control_frame, text="To:").grid(row=0, column=2, padx=(0, 5))
        self.end_date_var = tk.StringVar(value=date.today().strftime('%Y-%m-%d'))
        self.end_date_entry = ttk.Entry(control_frame, textvariable=self.end_date_var, width=12)
        self.end_date_entry.grid(row=0, column=3, padx=(0, 10))
        
        # Buttons
        ttk.Button(control_frame, text="Refresh Data", 
                  command=self.refresh_data).grid(row=0, column=4, padx=5)
        
        ttk.Button(control_frame, text="Export to Excel", 
                  command=self.export_to_excel).grid(row=0, column=5, padx=5)
        
        ttk.Button(control_frame, text="Generate Report", 
                  command=self.generate_analytics_report).grid(row=0, column=6, padx=5)
    
    def create_overview_tab(self):
        """Create overview tab with key metrics"""
        overview_frame = ttk.Frame(self.notebook)
        self.notebook.add(overview_frame, text="Overview")
        
        # Configure grid
        overview_frame.grid_rowconfigure(0, weight=1)
        overview_frame.grid_columnconfigure(0, weight=1)
        overview_frame.grid_columnconfigure(1, weight=1)
        
        # Left panel - KPIs
        kpi_frame = ttk.LabelFrame(overview_frame, text="Key Performance Indicators", padding="10")
        kpi_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=5)
        
        # KPI labels
        self.kpi_labels = {}
        kpi_metrics = [
            ("Total Revenue", "total_revenue"),
            ("Collected Amount", "collected_revenue"),
            ("Outstanding Amount", "outstanding_amount"),
            ("Collection Rate", "collection_rate"),
            ("Total Bills", "total_bills"),
            ("Active Patients", "active_patients")
        ]
        
        for i, (label, key) in enumerate(kpi_metrics):
            ttk.Label(kpi_frame, text=f"{label}:", font=('Arial', 10, 'bold')).grid(
                row=i, column=0, sticky="w", pady=2)
            self.kpi_labels[key] = ttk.Label(kpi_frame, text="₹0.00", font=('Arial', 10))
            self.kpi_labels[key].grid(row=i, column=1, sticky="e", pady=2, padx=(10, 0))
        
        # Right panel - Growth metrics
        growth_frame = ttk.LabelFrame(overview_frame, text="Month-over-Month Growth", padding="10")
        growth_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=5)
        
        # Growth chart placeholder
        self.growth_figure = Figure(figsize=(6, 4), dpi=80)
        self.growth_canvas = FigureCanvasTkAgg(self.growth_figure, growth_frame)
        self.growth_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        
        growth_frame.grid_rowconfigure(0, weight=1)
        growth_frame.grid_columnconfigure(0, weight=1)
    
    def create_revenue_tab(self):
        """Create revenue analytics tab"""
        revenue_frame = ttk.Frame(self.notebook)
        self.notebook.add(revenue_frame, text="Revenue")
        
        # Configure grid
        revenue_frame.grid_rowconfigure(0, weight=1)
        revenue_frame.grid_columnconfigure(0, weight=1)
        revenue_frame.grid_columnconfigure(1, weight=1)
        
        # Daily revenue trend
        daily_frame = ttk.LabelFrame(revenue_frame, text="Daily Revenue Trend", padding="10")
        daily_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=5)
        
        self.daily_figure = Figure(figsize=(6, 4), dpi=80)
        self.daily_canvas = FigureCanvasTkAgg(self.daily_figure, daily_frame)
        self.daily_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        
        daily_frame.grid_rowconfigure(0, weight=1)
        daily_frame.grid_columnconfigure(0, weight=1)
        
        # Payment method breakdown
        payment_frame = ttk.LabelFrame(revenue_frame, text="Payment Methods", padding="10")
        payment_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=5)
        
        self.payment_figure = Figure(figsize=(5, 4), dpi=80)
        self.payment_canvas = FigureCanvasTkAgg(self.payment_figure, payment_frame)
        self.payment_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        
        payment_frame.grid_rowconfigure(0, weight=1)
        payment_frame.grid_columnconfigure(0, weight=1)
    
    def create_tests_tab(self):
        """Create test analytics tab"""
        tests_frame = ttk.Frame(self.notebook)
        self.notebook.add(tests_frame, text="Tests")
        
        # Configure grid
        tests_frame.grid_rowconfigure(0, weight=1)
        tests_frame.grid_rowconfigure(1, weight=1)
        tests_frame.grid_columnconfigure(0, weight=1)
        tests_frame.grid_columnconfigure(1, weight=1)
        
        # Popular tests
        popular_frame = ttk.LabelFrame(tests_frame, text="Most Popular Tests", padding="10")
        popular_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=(0, 5))
        
        # Create treeview for popular tests
        popular_tree = ttk.Treeview(popular_frame, columns=('count', 'revenue'), show='tree headings', height=8)
        popular_tree.heading('#0', text='Test Name')
        popular_tree.heading('count', text='Count')
        popular_tree.heading('revenue', text='Revenue')
        popular_tree.column('#0', width=200)
        popular_tree.column('count', width=80, anchor='center')
        popular_tree.column('revenue', width=100, anchor='e')
        
        popular_scrollbar = ttk.Scrollbar(popular_frame, orient="vertical", command=popular_tree.yview)
        popular_tree.configure(yscrollcommand=popular_scrollbar.set)
        
        popular_tree.grid(row=0, column=0, sticky="nsew")
        popular_scrollbar.grid(row=0, column=1, sticky="ns")
        
        popular_frame.grid_rowconfigure(0, weight=1)
        popular_frame.grid_columnconfigure(0, weight=1)
        
        self.popular_tests_tree = popular_tree
        
        # Test results distribution
        results_frame = ttk.LabelFrame(tests_frame, text="Test Results Distribution", padding="10")
        results_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=(0, 5))
        
        self.results_figure = Figure(figsize=(5, 3), dpi=80)
        self.results_canvas = FigureCanvasTkAgg(self.results_figure, results_frame)
        self.results_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        
        results_frame.grid_rowconfigure(0, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)
        
        # Test volume by month
        volume_frame = ttk.LabelFrame(tests_frame, text="Test Volume Trend", padding="10")
        volume_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(5, 0))
        
        self.volume_figure = Figure(figsize=(10, 3), dpi=80)
        self.volume_canvas = FigureCanvasTkAgg(self.volume_figure, volume_frame)
        self.volume_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        
        volume_frame.grid_rowconfigure(0, weight=1)
        volume_frame.grid_columnconfigure(0, weight=1)
    
    def create_patients_tab(self):
        """Create patient analytics tab"""
        patients_frame = ttk.Frame(self.notebook)
        self.notebook.add(patients_frame, text="Patients")
        
        # Configure grid
        patients_frame.grid_rowconfigure(0, weight=1)
        patients_frame.grid_rowconfigure(1, weight=1)
        patients_frame.grid_columnconfigure(0, weight=1)
        patients_frame.grid_columnconfigure(1, weight=1)
        
        # Patient statistics
        stats_frame = ttk.LabelFrame(patients_frame, text="Patient Statistics", padding="10")
        stats_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=(0, 5))
        
        self.patient_stats_labels = {}
        patient_metrics = [
            ("New Patients", "new_patients"),
            ("Active Patients", "active_patients"),
            ("Repeat Patients", "repeat_patients"),
            ("Repeat Rate", "repeat_rate")
        ]
        
        for i, (label, key) in enumerate(patient_metrics):
            ttk.Label(stats_frame, text=f"{label}:", font=('Arial', 10, 'bold')).grid(
                row=i, column=0, sticky="w", pady=2)
            self.patient_stats_labels[key] = ttk.Label(stats_frame, text="0", font=('Arial', 10))
            self.patient_stats_labels[key].grid(row=i, column=1, sticky="e", pady=2, padx=(10, 0))
        
        # Gender distribution
        gender_frame = ttk.LabelFrame(patients_frame, text="Gender Distribution", padding="10")
        gender_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=(0, 5))
        
        self.gender_figure = Figure(figsize=(5, 3), dpi=80)
        self.gender_canvas = FigureCanvasTkAgg(self.gender_figure, gender_frame)
        self.gender_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        
        gender_frame.grid_rowconfigure(0, weight=1)
        gender_frame.grid_columnconfigure(0, weight=1)
        
        # Patient acquisition trend
        acquisition_frame = ttk.LabelFrame(patients_frame, text="Patient Acquisition Trend", padding="10")
        acquisition_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(5, 0))
        
        self.acquisition_figure = Figure(figsize=(10, 3), dpi=80)
        self.acquisition_canvas = FigureCanvasTkAgg(self.acquisition_figure, acquisition_frame)
        self.acquisition_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        
        acquisition_frame.grid_rowconfigure(0, weight=1)
        acquisition_frame.grid_columnconfigure(0, weight=1)
    
    def create_outstanding_tab(self):
        """Create outstanding payments tab"""
        outstanding_frame = ttk.Frame(self.notebook)
        self.notebook.add(outstanding_frame, text="Outstanding")
        
        # Configure grid
        outstanding_frame.grid_rowconfigure(0, weight=1)
        outstanding_frame.grid_rowconfigure(1, weight=1)
        outstanding_frame.grid_columnconfigure(0, weight=1)
        outstanding_frame.grid_columnconfigure(1, weight=1)
        
        # Aging analysis
        aging_frame = ttk.LabelFrame(outstanding_frame, text="Outstanding by Age", padding="10")
        aging_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=(0, 5))
        
        self.aging_figure = Figure(figsize=(5, 3), dpi=80)
        self.aging_canvas = FigureCanvasTkAgg(self.aging_figure, aging_frame)
        self.aging_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        
        aging_frame.grid_rowconfigure(0, weight=1)
        aging_frame.grid_columnconfigure(0, weight=1)
        
        # Top outstanding patients
        top_outstanding_frame = ttk.LabelFrame(outstanding_frame, text="Top Outstanding Patients", padding="10")
        top_outstanding_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=(0, 5))
        
        outstanding_tree = ttk.Treeview(top_outstanding_frame, columns=('phone', 'amount', 'bills'), 
                                       show='tree headings', height=8)
        outstanding_tree.heading('#0', text='Patient Name')
        outstanding_tree.heading('phone', text='Phone')
        outstanding_tree.heading('amount', text='Amount')
        outstanding_tree.heading('bills', text='Bills')
        outstanding_tree.column('#0', width=150)
        outstanding_tree.column('phone', width=100, anchor='center')
        outstanding_tree.column('amount', width=100, anchor='e')
        outstanding_tree.column('bills', width=60, anchor='center')
        
        outstanding_scrollbar = ttk.Scrollbar(top_outstanding_frame, orient="vertical", 
                                            command=outstanding_tree.yview)
        outstanding_tree.configure(yscrollcommand=outstanding_scrollbar.set)
        
        outstanding_tree.grid(row=0, column=0, sticky="nsew")
        outstanding_scrollbar.grid(row=0, column=1, sticky="ns")
        
        top_outstanding_frame.grid_rowconfigure(0, weight=1)
        top_outstanding_frame.grid_columnconfigure(0, weight=1)
        
        self.outstanding_tree = outstanding_tree
    
    def refresh_data(self):
        """Refresh all analytics data"""
        try:
            start_date = self.start_date_var.get()
            end_date = self.end_date_var.get()
            
            # Validate dates
            datetime.strptime(start_date, '%Y-%m-%d')
            datetime.strptime(end_date, '%Y-%m-%d')
            
            # Get analytics data
            revenue_data = self.analytics.get_revenue_summary(start_date, end_date)
            test_data = self.analytics.get_test_statistics(start_date, end_date)
            patient_data = self.analytics.get_patient_statistics(start_date, end_date)
            performance_data = self.analytics.get_performance_metrics()
            outstanding_data = self.analytics.get_outstanding_analysis()
            
            # Update all tabs
            self.update_overview_tab(revenue_data, patient_data, performance_data)
            self.update_revenue_tab(revenue_data)
            self.update_tests_tab(test_data)
            self.update_patients_tab(patient_data)
            self.update_outstanding_tab(outstanding_data)
            
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid date format. Use YYYY-MM-DD: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh data: {e}")
    
    def update_overview_tab(self, revenue_data, patient_data, performance_data):
        """Update overview tab with KPIs and growth chart"""
        # Update KPI labels
        self.kpi_labels['total_revenue'].config(text=f"₹{revenue_data['total_revenue']:,.2f}")
        self.kpi_labels['collected_revenue'].config(text=f"₹{revenue_data['collected_revenue']:,.2f}")
        self.kpi_labels['outstanding_amount'].config(text=f"₹{revenue_data['outstanding_amount']:,.2f}")
        self.kpi_labels['collection_rate'].config(text=f"{revenue_data['collection_rate']:.1f}%")
        self.kpi_labels['total_bills'].config(text=str(revenue_data['total_bills']))
        self.kpi_labels['active_patients'].config(text=str(patient_data['active_patients']))
        
        # Update growth chart
        self.growth_figure.clear()
        ax = self.growth_figure.add_subplot(111)
        
        metrics = ['Bills', 'Reports', 'Patients', 'Revenue']
        current = [
            performance_data['current_month']['bills'],
            performance_data['current_month']['reports'],
            performance_data['current_month']['patients'],
            performance_data['current_month']['revenue'] / 1000  # Show in thousands
        ]
        growth_rates = [
            performance_data['growth_rates']['bills'],
            performance_data['growth_rates']['reports'],
            performance_data['growth_rates']['patients'],
            performance_data['growth_rates']['revenue']
        ]
        
        colors = ['green' if rate >= 0 else 'red' for rate in growth_rates]
        bars = ax.bar(metrics, growth_rates, color=colors, alpha=0.7)
        
        ax.set_ylabel('Growth Rate (%)')
        ax.set_title('Month-over-Month Growth')
        ax.axhline(y=0, color='black', linestyle='-', alpha=0.3)
        
        # Add value labels on bars
        for bar, rate in zip(bars, growth_rates):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + (1 if height >= 0 else -3),
                   f'{rate:.1f}%', ha='center', va='bottom' if height >= 0 else 'top')
        
        self.growth_figure.tight_layout()
        self.growth_canvas.draw()
    
    def update_revenue_tab(self, revenue_data):
        """Update revenue tab with charts"""
        # Get daily revenue trend
        daily_data = self.analytics.get_daily_revenue_trend(30)
        
        # Daily revenue chart
        self.daily_figure.clear()
        ax = self.daily_figure.add_subplot(111)
        
        if daily_data:
            dates = [item['date'] for item in daily_data]
            revenues = [item['revenue'] for item in daily_data]
            collections = [item['collection'] for item in daily_data]
            
            ax.plot(dates, revenues, label='Revenue', marker='o', markersize=3)
            ax.plot(dates, collections, label='Collection', marker='s', markersize=3)
            ax.set_ylabel('Amount (₹)')
            ax.set_title('Daily Revenue Trend (Last 30 Days)')
            ax.legend()
            
            # Rotate x-axis labels
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        self.daily_figure.tight_layout()
        self.daily_canvas.draw()
        
        # Payment method pie chart
        self.payment_figure.clear()
        ax = self.payment_figure.add_subplot(111)
        
        payment_breakdown = revenue_data.get('payment_breakdown', {})
        if payment_breakdown:
            labels = list(payment_breakdown.keys())
            sizes = [payment_breakdown[label]['amount'] for label in labels]
            
            if sum(sizes) > 0:
                ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
                ax.set_title('Revenue by Payment Status')
        
        self.payment_canvas.draw()
    
    def update_tests_tab(self, test_data):
        """Update tests tab with popular tests and distribution"""
        # Clear and populate popular tests tree
        for item in self.popular_tests_tree.get_children():
            self.popular_tests_tree.delete(item)
        
        for test in test_data.get('popular_tests', []):
            self.popular_tests_tree.insert('', 'end', 
                                         text=f"{test['test_name']} ({test['test_code']})",
                                         values=(test['count'], f"₹{test['revenue']:.2f}"))
        
        # Test results distribution pie chart
        self.results_figure.clear()
        ax = self.results_figure.add_subplot(111)
        
        distribution = test_data.get('result_distribution', {})
        if any(distribution.values()):
            labels = []
            sizes = []
            colors = []
            
            if distribution['normal'] > 0:
                labels.append(f"Normal ({distribution['normal']})")
                sizes.append(distribution['normal'])
                colors.append('lightgreen')
            
            if distribution['abnormal'] > 0:
                labels.append(f"Abnormal ({distribution['abnormal']})")
                sizes.append(distribution['abnormal'])
                colors.append('lightcoral')
            
            if distribution['pending'] > 0:
                labels.append(f"Pending ({distribution['pending']})")
                sizes.append(distribution['pending'])
                colors.append('lightyellow')
            
            if sizes:
                ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
                ax.set_title('Test Results Distribution')
        
        self.results_canvas.draw()
    
    def update_patients_tab(self, patient_data):
        """Update patients tab with statistics and charts"""
        # Update patient statistics
        self.patient_stats_labels['new_patients'].config(text=str(patient_data['new_patients']))
        self.patient_stats_labels['active_patients'].config(text=str(patient_data['active_patients']))
        self.patient_stats_labels['repeat_patients'].config(text=str(patient_data['repeat_patients']))
        self.patient_stats_labels['repeat_rate'].config(text=f"{patient_data['repeat_rate']:.1f}%")
        
        # Gender distribution pie chart
        self.gender_figure.clear()
        ax = self.gender_figure.add_subplot(111)
        
        gender_dist = patient_data.get('gender_distribution', {})
        if gender_dist and any(gender_dist.values()):
            labels = list(gender_dist.keys())
            sizes = list(gender_dist.values())
            colors = ['lightblue', 'lightpink', 'lightgray'][:len(labels)]
            
            ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
            ax.set_title('Gender Distribution')
        
        self.gender_canvas.draw()
    
    def update_outstanding_tab(self, outstanding_data):
        """Update outstanding tab with aging and top patients"""
        # Aging analysis pie chart
        self.aging_figure.clear()
        ax = self.aging_figure.add_subplot(111)
        
        aging_analysis = outstanding_data.get('aging_analysis', [])
        if aging_analysis:
            labels = [item['period'] for item in aging_analysis if item['amount'] > 0]
            sizes = [item['amount'] for item in aging_analysis if item['amount'] > 0]
            
            if sizes:
                colors = ['lightgreen', 'yellow', 'orange', 'red'][:len(sizes)]
                ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
                ax.set_title('Outstanding by Age')
        
        self.aging_canvas.draw()
        
        # Clear and populate top outstanding patients
        for item in self.outstanding_tree.get_children():
            self.outstanding_tree.delete(item)
        
        for patient in outstanding_data.get('top_outstanding', []):
            self.outstanding_tree.insert('', 'end',
                                       text=patient['patient_name'],
                                       values=(patient['phone'], f"₹{patient['amount']:.2f}", patient['bills']))
    
    def export_to_excel(self):
        """Export analytics data to Excel"""
        try:
            from tkinter import filedialog
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Ask for save location
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Analytics Report"
            )
            
            if not filename:
                return
            
            # Get data
            start_date = self.start_date_var.get()
            end_date = self.end_date_var.get()
            
            revenue_data = self.analytics.get_revenue_summary(start_date, end_date)
            test_data = self.analytics.get_test_statistics(start_date, end_date)
            patient_data = self.analytics.get_patient_statistics(start_date, end_date)
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Revenue summary sheet
            ws_revenue = wb.active
            ws_revenue.title = "Revenue Summary"
            
            # Add headers
            headers = ["Metric", "Value"]
            for col, header in enumerate(headers, 1):
                cell = ws_revenue.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add revenue data
            revenue_rows = [
                ("Total Revenue", f"₹{revenue_data['total_revenue']:,.2f}"),
                ("Collected Revenue", f"₹{revenue_data['collected_revenue']:,.2f}"),
                ("Outstanding Amount", f"₹{revenue_data['outstanding_amount']:,.2f}"),
                ("Collection Rate", f"{revenue_data['collection_rate']:.2f}%"),
                ("Total Bills", revenue_data['total_bills'])
            ]
            
            for row, (metric, value) in enumerate(revenue_rows, 2):
                ws_revenue.cell(row=row, column=1, value=metric)
                ws_revenue.cell(row=row, column=2, value=value)
            
            # Test statistics sheet
            ws_tests = wb.create_sheet("Test Statistics")
            test_headers = ["Test Name", "Test Code", "Count", "Revenue"]
            
            for col, header in enumerate(test_headers, 1):
                cell = ws_tests.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row, test in enumerate(test_data.get('popular_tests', []), 2):
                ws_tests.cell(row=row, column=1, value=test['test_name'])
                ws_tests.cell(row=row, column=2, value=test['test_code'])
                ws_tests.cell(row=row, column=3, value=test['count'])
                ws_tests.cell(row=row, column=4, value=f"₹{test['revenue']:.2f}")
            
            # Save workbook
            wb.save(filename)
            messagebox.showinfo("Success", f"Analytics report exported to {filename}")
            
        except ImportError:
            messagebox.showerror("Error", "openpyxl is required for Excel export. Please install it.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {e}")
    
    def generate_analytics_report(self):
        """Generate comprehensive analytics PDF report"""
        try:
            from tkinter import filedialog
            from utils.pdf_generator import PDFGenerator, create_pdf_directory, generate_filename
            
            # Ask for save location
            pdf_dir = create_pdf_directory()
            filename = filedialog.asksaveasfilename(
                initialdir=pdf_dir,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Save Analytics Report"
            )
            
            if not filename:
                return
            
            # Get comprehensive analytics data
            start_date = self.start_date_var.get()
            end_date = self.end_date_var.get()
            
            analytics_data = {
                'report_period': f"{start_date} to {end_date}",
                'revenue': self.analytics.get_revenue_summary(start_date, end_date),
                'tests': self.analytics.get_test_statistics(start_date, end_date),
                'patients': self.analytics.get_patient_statistics(start_date, end_date),
                'performance': self.analytics.get_performance_metrics(),
                'outstanding': self.analytics.get_outstanding_analysis()
            }
            
            # Generate PDF (you would need to implement this method in PDFGenerator)
            pdf_gen = PDFGenerator()
            success = pdf_gen.generate_analytics_report_pdf(analytics_data, filename)
            
            if success:
                messagebox.showinfo("Success", f"Analytics report generated: {filename}")
            else:
                messagebox.showerror("Error", "Failed to generate PDF report")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {e}") 